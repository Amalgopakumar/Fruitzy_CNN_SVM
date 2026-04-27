"""
🍎🍌🥭 FruitAI Quality Classifier — Premium UI
===============================================
Modes:
  1. Single Image  — upload one image
  2. Folder Path   — type a local folder path
  3. ZIP Upload    — upload a zipped folder

Excel Report (auto-saved to same folder as app.py):
  Columns: Image Name | Fruit Type | Predicted Grade

Run:
    pip install streamlit pillow numpy tensorflow scikit-learn joblib openpyxl plotly
    streamlit run app.py
"""

import os
import io
import zipfile
import datetime
import streamlit as st
import numpy as np
import joblib
import tensorflow as tf
from PIL import Image
from pathlib import Path

# ─────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────

MODELS = {
    "🍎 Apple": {
        "cnn_path"      : "apple_cnn.h5",
        "svm_path"      : "apple_svm.pkl",
        "scaler_path"   : None,
        "feature_layer" : "feature_layer",
        "classes"       : ["Fresh", "Rotten"],
        "colors"        : ["#22c55e", "#ef4444"],
        "emoji"         : "🍎",
    },
    "🍌 Banana": {
        "cnn_path"      : "banana_cnn.h5",
        "svm_path"      : "banana_svm.pkl",
        "scaler_path"   : None,
        "feature_layer" : "dropout",
        "classes"       : ["Fresh", "Rotten", "Premium"],
        "colors"        : ["#22c55e", "#ef4444", "#f59e0b"],
        "emoji"         : "🍌",
    },
    "🥭 Mango": {
        "cnn_path"      : "mango_cnn.h5",
        "svm_path"      : "mango_svm.pkl",
        "scaler_path"   : "mango_scaler.pkl",
        "feature_layer" : "features",
        "classes"       : ["Good", "Premium", "Rotten"],
        "colors"        : ["#f59e0b", "#22c55e", "#ef4444"],
        "emoji"         : "🥭",
    },
}

IMG_SIZE    = (128, 128)
SUPPORTED   = {".jpg", ".jpeg", ".png", ".bmp", ".webp"}
EXCEL_PATH  = "fruit_predictions.xlsx"   # saved in same folder as app.py

# ─────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────

st.set_page_config(
    page_title = "FruitAI — Quality Classifier",
    page_icon  = "🍎",
    layout     = "wide",
)

# ─────────────────────────────────────────────────────────
# CSS + ANIMATIONS
# ─────────────────────────────────────────────────────────

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');

*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

html, body, [data-testid="stAppViewContainer"], [data-testid="stApp"] {
    font-family: 'Inter', sans-serif !important;
    background: transparent !important;
}

[data-testid="stApp"] {
    background: linear-gradient(135deg, #0f2027, #203a43, #2c5364) !important;
    min-height: 100vh;
}

.fruit-container {
    position: fixed; top: 0; left: 0;
    width: 100%; height: 100%;
    pointer-events: none; z-index: 0; overflow: hidden;
}
.fruit {
    position: absolute; top: -80px; font-size: 24px;
    animation: fall linear infinite; opacity: 0.25; user-select: none;
}
@keyframes fall {
    0%   { transform: translateY(-80px) rotate(0deg);   opacity: 0;    }
    10%  { opacity: 0.25; }
    90%  { opacity: 0.2;  }
    100% { transform: translateY(110vh) rotate(360deg); opacity: 0;    }
}
.fruit:nth-child(1)  { left:4%;  font-size:20px; animation-duration:8s;  animation-delay:0s;   }
.fruit:nth-child(2)  { left:10%; font-size:28px; animation-duration:11s; animation-delay:1s;   }
.fruit:nth-child(3)  { left:18%; font-size:16px; animation-duration:7s;  animation-delay:2.5s; }
.fruit:nth-child(4)  { left:26%; font-size:30px; animation-duration:13s; animation-delay:0.5s; }
.fruit:nth-child(5)  { left:34%; font-size:22px; animation-duration:9s;  animation-delay:3s;   }
.fruit:nth-child(6)  { left:42%; font-size:18px; animation-duration:10s; animation-delay:1.5s; }
.fruit:nth-child(7)  { left:50%; font-size:26px; animation-duration:12s; animation-delay:4s;   }
.fruit:nth-child(8)  { left:58%; font-size:20px; animation-duration:8s;  animation-delay:2s;   }
.fruit:nth-child(9)  { left:66%; font-size:32px; animation-duration:14s; animation-delay:0.8s; }
.fruit:nth-child(10) { left:74%; font-size:18px; animation-duration:9s;  animation-delay:3.5s; }
.fruit:nth-child(11) { left:82%; font-size:24px; animation-duration:11s; animation-delay:1.2s; }
.fruit:nth-child(12) { left:90%; font-size:20px; animation-duration:7s;  animation-delay:4.5s; }
.fruit:nth-child(13) { left:96%; font-size:28px; animation-duration:10s; animation-delay:2.8s; }
.fruit:nth-child(14) { left:8%;  font-size:14px; animation-duration:6s;  animation-delay:5s;   }
.fruit:nth-child(15) { left:55%; font-size:16px; animation-duration:15s; animation-delay:0.2s; }

[data-testid="stMainBlockContainer"] {
    position: relative; z-index: 1;
    max-width: 1200px !important;
    margin: 0 auto !important;
    padding: 2rem 1.5rem !important;
}

.app-header {
    text-align: center; padding: 2.5rem 2rem 2rem;
    background: rgba(255,255,255,0.04);
    border: 1px solid rgba(255,255,255,0.1);
    border-radius: 24px; backdrop-filter: blur(20px);
    margin-bottom: 2rem; box-shadow: 0 8px 32px rgba(0,0,0,0.3);
}
.app-header h1 {
    font-size: 3rem; font-weight: 900;
    background: linear-gradient(135deg, #fff 0%, #a8edea 50%, #fed6e3 100%);
    -webkit-background-clip: text; -webkit-text-fill-color: transparent;
    background-clip: text; letter-spacing: -1px; line-height: 1.1;
}
.app-header p { color: rgba(255,255,255,0.6); font-size:1.05rem; margin-top:0.6rem; }
.header-fruits {
    font-size: 2rem; letter-spacing: 6px; margin-bottom: 1rem;
    animation: pulse 2s ease-in-out infinite;
}
@keyframes pulse {
    0%,100% { transform: scale(1); }
    50%      { transform: scale(1.08); }
}

.glass-card {
    background: rgba(255,255,255,0.07);
    border: 1px solid rgba(255,255,255,0.12);
    border-radius: 20px; backdrop-filter: blur(20px);
    padding: 2rem; box-shadow: 0 8px 32px rgba(0,0,0,0.25);
}
.glass-card h3 {
    color: rgba(255,255,255,0.9); font-size: 1rem; font-weight: 700;
    text-transform: uppercase; letter-spacing: 2px; margin-bottom: 1.5rem;
}

/* Tab styling */
[data-testid="stTabs"] [data-testid="stMarkdownContainer"] p {
    color: white !important;
}
.stTabs [data-baseweb="tab-list"] {
    background: rgba(255,255,255,0.06) !important;
    border-radius: 12px !important; padding: 4px !important;
    gap: 4px !important;
}
.stTabs [data-baseweb="tab"] {
    background: transparent !important;
    color: rgba(255,255,255,0.5) !important;
    border-radius: 8px !important;
    font-weight: 600 !important; font-size: 0.85rem !important;
}
.stTabs [aria-selected="true"] {
    background: rgba(255,255,255,0.15) !important;
    color: white !important;
}

[data-testid="stSelectbox"] label,
[data-testid="stFileUploader"] label,
[data-testid="stTextInput"] label {
    color: rgba(255,255,255,0.7) !important;
    font-weight: 600 !important; font-size: 0.85rem !important;
    text-transform: uppercase !important; letter-spacing: 1px !important;
}
[data-testid="stSelectbox"] > div > div {
    background: rgba(255,255,255,0.08) !important;
    border: 1px solid rgba(255,255,255,0.2) !important;
    border-radius: 12px !important; color: white !important;
}
[data-testid="stTextInput"] input {
    background: rgba(255,255,255,0.08) !important;
    border: 1px solid rgba(255,255,255,0.2) !important;
    border-radius: 12px !important; color: white !important;
}
[data-testid="stFileUploader"] {
    background: rgba(255,255,255,0.04) !important;
    border: 2px dashed rgba(255,255,255,0.2) !important;
    border-radius: 16px !important; padding: 1rem !important;
}

[data-testid="stButton"] > button {
    width: 100% !important;
    background: linear-gradient(135deg, #11998e, #38ef7d) !important;
    color: white !important; border: none !important;
    border-radius: 14px !important; padding: 0.85rem 2rem !important;
    font-size: 1rem !important; font-weight: 700 !important;
    transition: all 0.3s ease !important;
    box-shadow: 0 4px 20px rgba(56,239,125,0.3) !important;
    margin-top: 0.5rem !important;
}
[data-testid="stButton"] > button:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 30px rgba(56,239,125,0.5) !important;
}
[data-testid="stButton"] > button:disabled {
    background: rgba(255,255,255,0.1) !important;
    box-shadow: none !important; transform: none !important;
    color: rgba(255,255,255,0.3) !important;
}

[data-testid="stImage"] img {
    border-radius: 14px !important;
    border: 2px solid rgba(255,255,255,0.15) !important;
    box-shadow: 0 4px 20px rgba(0,0,0,0.3) !important;
    width: 100% !important;
}

.result-card {
    background: rgba(255,255,255,0.06);
    border: 1px solid rgba(255,255,255,0.12);
    border-radius: 20px; padding: 2rem;
    animation: fadeSlideIn 0.5s ease forwards;
    margin-bottom: 1rem;
}
@keyframes fadeSlideIn {
    from { opacity:0; transform:translateY(20px); }
    to   { opacity:1; transform:translateY(0);    }
}
.result-label {
    font-size: 0.75rem; font-weight: 700;
    text-transform: uppercase; letter-spacing: 2px;
    color: rgba(255,255,255,0.45); margin-bottom: 0.4rem;
}
.fruit-name-display {
    font-size: 1.3rem; font-weight: 800;
    color: white; margin-bottom: 1.5rem;
}
.grade-pill {
    display: inline-flex; align-items: center; gap: 10px;
    padding: 12px 28px; border-radius: 50px;
    font-size: 1.5rem; font-weight: 900; letter-spacing: 1px;
    margin-bottom: 0.5rem;
    animation: popIn 0.4s cubic-bezier(0.175,0.885,0.32,1.275) forwards;
}
@keyframes popIn {
    from { transform:scale(0.7); opacity:0; }
    to   { transform:scale(1);   opacity:1; }
}
.confidence-text {
    font-size: 0.85rem; color: rgba(255,255,255,0.5);
    margin-top: 0.5rem; margin-bottom: 1rem;
}
.prob-row { margin: 8px 0; }
.prob-header {
    display: flex; justify-content: space-between;
    font-size: 0.82rem; font-weight: 600;
    color: rgba(255,255,255,0.75); margin-bottom: 4px;
}
.prob-track {
    background: rgba(255,255,255,0.08);
    border-radius: 10px; height: 10px; overflow: hidden;
}
.prob-fill { height: 10px; border-radius: 10px; }

.verdict-banner {
    border-radius: 16px; padding: 1rem 1.4rem;
    font-size: 0.95rem; font-weight: 600;
    display: flex; align-items: center; gap: 10px;
    margin-top: 1rem;
    animation: fadeSlideIn 0.6s ease 0.3s forwards; opacity:0;
}
.verdict-fresh, .verdict-premium {
    background: rgba(34,197,94,0.15);
    border: 1px solid rgba(34,197,94,0.35); color: #86efac;
}
.verdict-good {
    background: rgba(245,158,11,0.15);
    border: 1px solid rgba(245,158,11,0.35); color: #fcd34d;
}
.verdict-rotten {
    background: rgba(239,68,68,0.15);
    border: 1px solid rgba(239,68,68,0.35); color: #fca5a5;
}

/* Batch results table */
.batch-row {
    display: flex; align-items: center; gap: 12px;
    padding: 10px 14px; margin: 6px 0;
    background: rgba(255,255,255,0.05);
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 10px; font-size: 0.88rem;
}
.batch-name { flex:1; color: rgba(255,255,255,0.8); font-weight:500; }
.batch-grade { font-weight:700; padding: 3px 12px; border-radius:20px; font-size:0.82rem; }

/* Excel saved banner */
.excel-banner {
    background: rgba(34,197,94,0.12);
    border: 1px solid rgba(34,197,94,0.3);
    border-radius: 12px; padding: 12px 16px;
    color: #86efac; font-size: 0.88rem;
    display: flex; align-items: center; gap: 8px;
    margin-top: 1rem;
}

.placeholder-box {
    display: flex; flex-direction: column;
    align-items: center; justify-content: center;
    height: 260px; color: rgba(255,255,255,0.25);
    font-size: 0.95rem; gap: 1rem;
    border: 2px dashed rgba(255,255,255,0.1);
    border-radius: 16px;
}
.placeholder-box .big-icon { font-size: 3.5rem; opacity: 0.4; }

#MainMenu, footer, header { visibility: hidden; }
[data-testid="stToolbar"] { display: none; }
.footer {
    text-align: center; color: rgba(255,255,255,0.2);
    font-size: 0.78rem; padding: 2rem 0 1rem; letter-spacing: 1px;
}
</style>

<!-- Falling Fruits -->
<div class="fruit-container">
    <div class="fruit">🍎</div><div class="fruit">🍌</div>
    <div class="fruit">🥭</div><div class="fruit">🍊</div>
    <div class="fruit">🍇</div><div class="fruit">🍓</div>
    <div class="fruit">🍋</div><div class="fruit">🍎</div>
    <div class="fruit">🥭</div><div class="fruit">🍌</div>
    <div class="fruit">🍊</div><div class="fruit">🍓</div>
    <div class="fruit">🍎</div><div class="fruit">🍇</div>
    <div class="fruit">🍋</div>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────────────────

@st.cache_resource
def load_models(cnn_path, svm_path, scaler_path, feature_layer):
    cnn = tf.keras.models.load_model(cnn_path)
    fe  = tf.keras.Model(inputs=cnn.input,
                         outputs=cnn.get_layer(feature_layer).output)
    svm    = joblib.load(svm_path)
    scaler = joblib.load(scaler_path) if scaler_path else None
    return fe, svm, scaler


def preprocess_image(img):
    img = img.convert("RGB").resize(IMG_SIZE)
    arr = np.array(img, dtype=np.float32) / 255.0
    return np.expand_dims(arr, axis=0)


def predict(fe, svm, scaler, img_array, classes):
    features = fe.predict(img_array, verbose=0)
    if scaler is not None:
        features = scaler.transform(features)
    raw               = svm.predict(features)[0]
    svm_lower         = [str(s).lower() for s in svm.classes_]
    raw_lower         = str(raw).lower()
    class_idx         = svm_lower.index(raw_lower) if raw_lower in svm_lower else int(raw)
    class_idx         = min(class_idx, len(classes) - 1)
    label             = classes[class_idx]
    try:
        probs      = svm.predict_proba(features)[0]
        confidence = float(probs[class_idx]) * 100
    except AttributeError:
        probs            = np.zeros(len(classes))
        probs[class_idx] = 1.0
        confidence       = 100.0
    return label, confidence, probs


def grade_icon(l):
    return {"premium":"👑","fresh":"✅","good":"👍","rotten":"❌"}.get(l,"🔍")


def verdict_html(ll, label):
    v = {
        "premium": ("verdict-fresh verdict-premium","👑","Premium Quality — Excellent! Ready for top-tier market."),
        "fresh"  : ("verdict-fresh","✅","Fresh — Great condition! Ready for sale."),
        "good"   : ("verdict-good","👍","Good Quality — Acceptable. Suitable for regular market."),
        "rotten" : ("verdict-rotten","❌","Rotten — Not suitable for consumption."),
    }
    cls, icon, msg = v.get(ll, ("verdict-good","ℹ️",f"Predicted: {label}"))
    return f'<div class="verdict-banner {cls}">{icon} {msg}</div>'


def render_radar(cfg, probs, grade_color):
    import plotly.graph_objects as go
    labels = [c.capitalize() for c in cfg["classes"]]
    values = [float(p) * 100 for p in probs]
    lc     = labels + [labels[0]]
    vc     = values + [values[0]]
    r, g, b = int(grade_color[1:3],16), int(grade_color[3:5],16), int(grade_color[5:7],16)
    fig = go.Figure()
    fig.add_trace(go.Scatterpolar(
        r=vc, theta=lc, fill="toself",
        fillcolor=f"rgba({r},{g},{b},0.25)",
        line=dict(color=grade_color, width=2.5),
        mode="lines+markers",
        marker=dict(size=8, color=grade_color, line=dict(color="white",width=1.5)),
    ))
    fig.update_layout(
        polar=dict(
            bgcolor="rgba(255,255,255,0.04)",
            radialaxis=dict(visible=True, range=[0,100], ticksuffix="%",
                tickfont=dict(size=9,color="rgba(255,255,255,0.4)"),
                gridcolor="rgba(255,255,255,0.1)", linecolor="rgba(255,255,255,0.1)"),
            angularaxis=dict(tickfont=dict(size=13,color="white",family="Inter"),
                gridcolor="rgba(255,255,255,0.1)", linecolor="rgba(255,255,255,0.15)"),
        ),
        paper_bgcolor="rgba(0,0,0,0)", font=dict(color="white",family="Inter"),
        margin=dict(t=50,b=50,l=60,r=60), height=300, showlegend=False,
        title=dict(text="Grade Probability Radar",
                   font=dict(size=12,color="rgba(255,255,255,0.45)"),x=0.5),
    )
    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar":False})


def save_to_excel(records):
    """
    Saves prediction records to Excel in same folder as app.py.
    Appends to existing file if it exists.
    records = list of dicts: {image_name, fruit_type, predicted_grade}
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    excel_path = Path(__file__).parent / EXCEL_PATH

    # Load existing or create new
    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fruit Predictions"

        # Header row styling
        headers = ["#", "Image Name", "Fruit Type", "Predicted Grade", "Date & Time"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.fill      = PatternFill("solid", fgColor="1A3C5E")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Column widths
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 22
        ws.row_dimensions[1].height = 22

    # Grade colors for Excel
    grade_colors = {
        "fresh"  : "22C55E", "premium": "22C55E",
        "good"   : "F59E0B", "rotten" : "EF4444",
    }

    start_row = ws.max_row + 1
    now       = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for i, rec in enumerate(records):
        row = start_row + i
        sn  = ws.cell(row=row, column=1, value=ws.max_row)
        sn.alignment = Alignment(horizontal="center")

        ws.cell(row=row, column=2, value=rec["image_name"])
        ws.cell(row=row, column=3, value=rec["fruit_type"])

        grade_cell       = ws.cell(row=row, column=4, value=rec["predicted_grade"])
        grade_cell.font  = Font(bold=True, color="FFFFFF")
        clr              = grade_colors.get(rec["predicted_grade"].lower(), "888888")
        grade_cell.fill  = PatternFill("solid", fgColor=clr)
        grade_cell.alignment = Alignment(horizontal="center")

        ws.cell(row=row, column=5, value=now)

        # Alternating row color
        bg = "F0F4F8" if i % 2 == 0 else "FFFFFF"
        for col in [1, 2, 3, 5]:
            c = ws.cell(row=row, column=col)
            c.fill = PatternFill("solid", fgColor=bg)

    # Add summary row at bottom
    total_row = ws.max_row + 2
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=2, value=f"{len(records)} images processed")
    for col in range(1, 6):
        c = ws.cell(row=total_row, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A3C5E")

    wb.save(excel_path)
    return excel_path


def render_single_result(cfg, label, confidence, probs, fruit_name):
    """Renders the full result card + prob bars + radar for single image."""
    grade_color = cfg["colors"][cfg["classes"].index(label)]
    label_lower = label.lower()
    icon        = grade_icon(label_lower)

    st.markdown(f"""
    <div class="result-card">
        <div class="result-label">Fruit</div>
        <div class="fruit-name-display">{cfg['emoji']} {fruit_name}</div>
        <div class="result-label">Predicted Grade</div>
        <div class="grade-pill"
             style="background:{grade_color}22; color:{grade_color};
                    border:2px solid {grade_color}55;
                    box-shadow:0 0 24px {grade_color}33;">
            <span>{icon}</span> {label.capitalize()}
        </div>
        <div class="confidence-text">
            Model confidence: <strong style="color:{grade_color}">{confidence:.1f}%</strong>
        </div>
        <div class="result-label" style="margin-top:16px; margin-bottom:10px;">
            Class Probabilities
        </div>
    </div>
    """, unsafe_allow_html=True)

    for c, p, clr in zip(cfg["classes"], probs, cfg["colors"]):
        pct = float(p) * 100
        st.markdown(f"""
        <div class="prob-row">
            <div class="prob-header">
                <span>{c.capitalize()}</span>
                <span style="color:{clr}; font-weight:700">{pct:.1f}%</span>
            </div>
            <div class="prob-track">
                <div class="prob-fill"
                     style="width:{pct:.1f}%; background:{clr};
                            box-shadow:0 0 8px {clr}88;"></div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown(verdict_html(label_lower, label), unsafe_allow_html=True)
    render_radar(cfg, probs, grade_color)


# ─────────────────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────────────────

st.markdown("""
<div class="app-header">
    <div class="header-fruits">🍎 🍌 🥭</div>
    <h1>FruitAI Quality Classifier</h1>
    <p>Premium AI-powered fruit quality grading — CNN + SVM</p>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────
# MAIN LAYOUT
# ─────────────────────────────────────────────────────────

col_left, col_right = st.columns([1, 1], gap="large")

# ══════════════════════════════════════════════════════════
# LEFT PANEL — Input
# ══════════════════════════════════════════════════════════
with col_left:
    st.markdown('<div class="glass-card">', unsafe_allow_html=True)
    st.markdown('<h3>🔍 Input</h3>', unsafe_allow_html=True)

    fruit_choice = st.selectbox("Select Fruit", options=list(MODELS.keys()))
    cfg          = MODELS[fruit_choice]
    fruit_name   = fruit_choice.split(" ", 1)[1]

    # ── Mode Tabs ──
    tab_single, tab_folder, tab_zip = st.tabs([
        "📸 Single Image", "📁 Folder Path", "🗜️ ZIP Upload"
    ])

    # ── Tab 1: Single Image ──
    with tab_single:
        uploaded_file = st.file_uploader(
            "Upload Fruit Image",
            type=["jpg","jpeg","png","bmp","webp"],
            key="single"
        )
        if uploaded_file:
            st.image(Image.open(uploaded_file),
                     caption=f"📸 {uploaded_file.name}",
                     use_container_width=True)

        btn_single = st.button(
            f"🔬 Analyse {cfg['emoji']}",
            key="btn_single",
            use_container_width=True,
            disabled=(uploaded_file is None)
        )

    # ── Tab 2: Folder Path ──
    with tab_folder:
        st.markdown(
            '<p style="color:rgba(255,255,255,0.5); font-size:0.8rem; margin-bottom:8px;">'
            '⚠️ Only works when app runs on the same machine as your files</p>',
            unsafe_allow_html=True
        )
        folder_path = st.text_input(
            "Enter Folder Path",
            placeholder=r"C:/Fruit_Fresh_pro/test/mango/",
            key="folder_path"
        )

        folder_images = []
        if folder_path and os.path.isdir(folder_path):
            folder_images = [
                f for f in os.listdir(folder_path)
                if Path(f).suffix.lower() in SUPPORTED
            ]
            st.markdown(
                f'<p style="color:#86efac; font-size:0.85rem;">✅ Found {len(folder_images)} images</p>',
                unsafe_allow_html=True
            )
        elif folder_path:
            st.markdown(
                '<p style="color:#fca5a5; font-size:0.85rem;">❌ Folder not found</p>',
                unsafe_allow_html=True
            )

        btn_folder = st.button(
            f"🔬 Analyse Folder {cfg['emoji']}",
            key="btn_folder",
            use_container_width=True,
            disabled=(len(folder_images) == 0)
        )

    # ── Tab 3: ZIP Upload ──
    with tab_zip:
        uploaded_zip = st.file_uploader(
            "Upload Folder as ZIP",
            type=["zip"],
            key="zip"
        )

        zip_images = []
        if uploaded_zip:
            with zipfile.ZipFile(io.BytesIO(uploaded_zip.read())) as z:
                zip_images = [
                    f for f in z.namelist()
                    if Path(f).suffix.lower() in SUPPORTED
                    and not f.startswith("__MACOSX")
                ]
            st.markdown(
                f'<p style="color:#86efac; font-size:0.85rem;">✅ Found {len(zip_images)} images in ZIP</p>',
                unsafe_allow_html=True
            )

        btn_zip = st.button(
            f"🔬 Analyse ZIP {cfg['emoji']}",
            key="btn_zip",
            use_container_width=True,
            disabled=(len(zip_images) == 0)
        )

    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════
# RIGHT PANEL — Output
# ══════════════════════════════════════════════════════════
with col_right:
    st.markdown('<div class="glass-card">', unsafe_allow_html=True)
    st.markdown('<h3>📊 Result</h3>', unsafe_allow_html=True)

    # ── Load models (shared for all modes) ──
    def get_models():
        try:
            return load_models(cfg["cnn_path"], cfg["svm_path"],
                               cfg["scaler_path"], cfg["feature_layer"])
        except FileNotFoundError as e:
            st.error(f"❌ File not found: {e}")
            st.stop()
        except Exception as e:
            st.error(f"❌ Error loading model: {e}")
            st.stop()

    # ────────────────────────────────────────
    # MODE 1: Single Image
    # ────────────────────────────────────────
    if btn_single and uploaded_file:
        with st.spinner("Analysing..."):
            fe, svm, scaler = get_models()
            img_array       = preprocess_image(Image.open(uploaded_file))
            label, conf, probs = predict(fe, svm, scaler, img_array, cfg["classes"])

        render_single_result(cfg, label, conf, probs, fruit_name)

        # Save to Excel
        save_to_excel([{
            "image_name"     : uploaded_file.name,
            "fruit_type"     : fruit_name,
            "predicted_grade": label,
        }])
        excel_path = Path(__file__).parent / EXCEL_PATH
        st.markdown(
            f'<div class="excel-banner">📊 Result saved to Excel → <strong>{EXCEL_PATH}</strong></div>',
            unsafe_allow_html=True
        )

    # ────────────────────────────────────────
    # MODE 2: Folder Path
    # ────────────────────────────────────────
    elif btn_folder and folder_images:
        fe, svm, scaler = get_models()
        records         = []
        results_html    = ""

        progress = st.progress(0, text="Analysing images...")
        for i, fname in enumerate(folder_images):
            img_path = os.path.join(folder_path, fname)
            try:
                img        = Image.open(img_path)
                arr        = preprocess_image(img)
                label, conf, probs = predict(fe, svm, scaler, arr, cfg["classes"])
                records.append({
                    "image_name"     : fname,
                    "fruit_type"     : fruit_name,
                    "predicted_grade": label,
                })
                clr = cfg["colors"][cfg["classes"].index(label)]
                results_html += f"""
                <div class="batch-row">
                    <span class="batch-name">📄 {fname}</span>
                    <span class="batch-grade"
                          style="background:{clr}22; color:{clr}; border:1px solid {clr}55;">
                        {grade_icon(label.lower())} {label.capitalize()}
                    </span>
                </div>"""
            except Exception as e:
                results_html += f'<div class="batch-row"><span class="batch-name">⚠️ {fname} — Error: {e}</span></div>'

            progress.progress((i + 1) / len(folder_images),
                              text=f"Processing {i+1}/{len(folder_images)}...")

        progress.empty()

        # Summary card
        st.markdown(f"""
        <div class="result-card">
            <div class="result-label">Batch Complete</div>
            <div class="fruit-name-display">{cfg['emoji']} {fruit_name} — {len(records)} images</div>
        </div>
        """, unsafe_allow_html=True)

        # Batch rows rendered one by one
        for rec in records:
            clr = cfg["colors"][cfg["classes"].index(rec["predicted_grade"])]
            icon = grade_icon(rec["predicted_grade"].lower())
            st.markdown(f"""
            <div class="batch-row">
                <span class="batch-name">📄 {rec['image_name']}</span>
                <span class="batch-grade"
                      style="background:{clr}22; color:{clr}; border:1px solid {clr}55;">
                    {icon} {rec['predicted_grade'].capitalize()}
                </span>
            </div>
            """, unsafe_allow_html=True)

        # Grade distribution pie chart
        if records:
            import plotly.graph_objects as go
            from collections import Counter
            grade_counts = Counter(r["predicted_grade"] for r in records)
            labels = list(grade_counts.keys())
            values = list(grade_counts.values())
            colors = [cfg["colors"][cfg["classes"].index(l)] if l in cfg["classes"] else "#888" for l in labels]

            fig = go.Figure(go.Pie(
                labels   = [l.capitalize() for l in labels],
                values   = values,
                marker   = dict(colors=colors, line=dict(color="rgba(255,255,255,0.2)", width=2)),
                textinfo = "label+percent+value",
                textfont = dict(size=13, color="white"),
                hole     = 0.4,
            ))
            fig.update_layout(
                paper_bgcolor = "rgba(0,0,0,0)",
                font          = dict(color="white", family="Inter"),
                margin        = dict(t=50, b=20, l=20, r=20),
                height        = 300,
                showlegend    = True,
                legend        = dict(font=dict(color="white"), bgcolor="rgba(0,0,0,0)"),
                title         = dict(
                    text = f"Grade Distribution ({len(records)} images)",
                    font = dict(size=13, color="rgba(255,255,255,0.5)"),
                    x    = 0.5,
                ),
            )
            st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

        # Save to Excel
        if records:
            save_to_excel(records)
            st.markdown(
                f'<div class="excel-banner">📊 {len(records)} results saved to Excel → <strong>{EXCEL_PATH}</strong></div>',
                unsafe_allow_html=True
            )

    # ────────────────────────────────────────
    # MODE 3: ZIP Upload
    # ────────────────────────────────────────
    elif btn_zip and zip_images and uploaded_zip:
        fe, svm, scaler = get_models()
        records         = []
        results_html    = ""

        # Re-read zip (after previous read)
        uploaded_zip.seek(0)
        with zipfile.ZipFile(io.BytesIO(uploaded_zip.read())) as z:
            progress = st.progress(0, text="Analysing ZIP images...")
            for i, fname in enumerate(zip_images):
                try:
                    with z.open(fname) as f:
                        img   = Image.open(io.BytesIO(f.read()))
                        arr   = preprocess_image(img)
                        label, conf, probs = predict(fe, svm, scaler, arr, cfg["classes"])
                        name  = Path(fname).name
                        records.append({
                            "image_name"     : name,
                            "fruit_type"     : fruit_name,
                            "predicted_grade": label,
                        })
                        clr = cfg["colors"][cfg["classes"].index(label)]
                        results_html += f"""
                        <div class="batch-row">
                            <span class="batch-name">📄 {name}</span>
                            <span class="batch-grade"
                                  style="background:{clr}22; color:{clr}; border:1px solid {clr}55;">
                                {grade_icon(label.lower())} {label.capitalize()}
                            </span>
                        </div>"""
                except Exception as e:
                    results_html += f'<div class="batch-row"><span class="batch-name">⚠️ {Path(fname).name} — Error: {e}</span></div>'

                progress.progress((i + 1) / len(zip_images),
                                  text=f"Processing {i+1}/{len(zip_images)}...")

            progress.empty()

        st.markdown(f"""
        <div class="result-card">
            <div class="result-label">ZIP Batch Complete</div>
            <div class="fruit-name-display">{cfg['emoji']} {fruit_name} — {len(records)} images</div>
        </div>
        """, unsafe_allow_html=True)

        # Batch rows rendered one by one
        for rec in records:
            clr  = cfg["colors"][cfg["classes"].index(rec["predicted_grade"])]
            icon = grade_icon(rec["predicted_grade"].lower())
            st.markdown(f"""
            <div class="batch-row">
                <span class="batch-name">📄 {rec['image_name']}</span>
                <span class="batch-grade"
                      style="background:{clr}22; color:{clr}; border:1px solid {clr}55;">
                    {icon} {rec['predicted_grade'].capitalize()}
                </span>
            </div>
            """, unsafe_allow_html=True)

        # Grade distribution pie chart
        if records:
            import plotly.graph_objects as go
            from collections import Counter
            grade_counts = Counter(r["predicted_grade"] for r in records)
            labels = list(grade_counts.keys())
            values = list(grade_counts.values())
            colors = [cfg["colors"][cfg["classes"].index(l)] if l in cfg["classes"] else "#888" for l in labels]

            fig = go.Figure(go.Pie(
                labels   = [l.capitalize() for l in labels],
                values   = values,
                marker   = dict(colors=colors, line=dict(color="rgba(255,255,255,0.2)", width=2)),
                textinfo = "label+percent+value",
                textfont = dict(size=13, color="white"),
                hole     = 0.4,
            ))
            fig.update_layout(
                paper_bgcolor = "rgba(0,0,0,0)",
                font          = dict(color="white", family="Inter"),
                margin        = dict(t=50, b=20, l=20, r=20),
                height        = 300,
                showlegend    = True,
                legend        = dict(font=dict(color="white"), bgcolor="rgba(0,0,0,0)"),
                title         = dict(
                    text = f"Grade Distribution ({len(records)} images)",
                    font = dict(size=13, color="rgba(255,255,255,0.5)"),
                    x    = 0.5,
                ),
            )
            st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

        if records:
            save_to_excel(records)
            st.markdown(
                f'<div class="excel-banner">📊 {len(records)} results saved to Excel → <strong>{EXCEL_PATH}</strong></div>',
                unsafe_allow_html=True
            )

    # ── Placeholder ──
    else:
        st.markdown("""
        <div class="placeholder-box">
            <div class="big-icon">🍽️</div>
            <div>Choose a mode and upload to begin</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────
# FOOTER
# ─────────────────────────────────────────────────────────

st.markdown("""
<div class="footer">
    FruitAI Quality Classifier &nbsp;·&nbsp; CNN + SVM &nbsp;·&nbsp; Built with Streamlit
</div>
""", unsafe_allow_html=True)
